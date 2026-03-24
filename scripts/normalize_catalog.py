import json, re, sys
import openpyxl
from collections import defaultdict

PROV_NORM = {
    'Aseo': 'Aseo',
    'Cafe Caribe': 'Café Caribe',
    'Cafeteria': 'Cafetería',
    'Coporate Coffe': 'Corporate Coffee',
    'Corporate Coffe': 'Corporate Coffee',
    'Papeleria': 'Papelería',
}
# latin-1 variant of Papelería
PROV_NORM['Papelar\xeda'] = 'Papelería'

PROD_NORM = {
    'Aceite 500cc':'Aceite de Oliva','Aceite de Oliva':'Aceite de Oliva','Aceite de oliva':'Aceite de Oliva',
    'Aceto':'Aceto Balsámico','Aceto 500cc':'Aceto Balsámico','Aceto Balsamico':'Aceto Balsámico',
    'Aerosol':'Ambientador','Aerosol de aroma':'Ambientador','Ambientador':'Ambientador',
    'Desodorante Ambiental':'Ambientador','Desodorante ambiental':'Ambientador',
    'Antigrasa x 5L':'Antigrasa','Limpiador Antigrasa':'Antigrasa',
    'Antisarro x 5 L':'Antisarro',
    'Atomizador Pistola':'Atomizador',
    'Azucar':'Azúcar','Azucar (UN)':'Azúcar','Azucar (caja)':'Azúcar','Azucar kg':'Azúcar',
    'Sobre Azucar x Caja (800 S)':'Azúcar',
    'Bolsas Basura':'Bolsas Basura','Bolsas Basura  70x90':'Bolsas 70x90','Bolsas Basura  80x110':'Bolsas 80x110',
    'Bolsas Basura 50x70':'Bolsas 50x70','Bolsas Basura 70x90':'Bolsas 70x90','Bolsas Basura 80x110':'Bolsas 80x110',
    'Bolsas L':'Bolsas L','Bolsas M':'Bolsas M','Bolsas S':'Bolsas S',
    'Cafe Caribe Grano (KG)':'Café Grano (KG)','Cafe Grano (KG)':'Café Grano (KG)',
    'Cafe Granulado':'Café Granulado',
    'CIF':'CIF','Cif en crema':'CIF',
    'Cloro Gel':'Cloro Gel',
    'Chocolate':'Chocolate','Chocolate (KG)':'Chocolate',
    'Endulzante':'Endulzante','Sucralosa (UN)':'Endulzante',
    'Esponja 4U':'Esponja',
    'Guantes':'Guantes','Guantes Multiuso':'Guantes',
    'Igenix':'Igenix',
    'Jabon Liquido x 5L':'Jabón de Mano','Jabon liquido':'Jabón de Mano','Repuesto Jabon':'Repuesto Jabón',
    'Lavaloza':'Lavaloza','Lavalozas':'Lavaloza','Lavavajillas':'Lavaloza','Jabon Lavaloza x 5L':'Lavaloza',
    'Leche Liquida':'Leche Líquida','Leche liquida':'Leche Líquida',
    'Leche en polvo':'Leche en Polvo','Leche en polvo (KG)':'Leche en Polvo',
    'Jugo de Limon':'Limón','Limon 1 Lt':'Limón','Sucedaneo Limon':'Limón',
    'Limpia Piso':'Limpia Piso','Limpia piso':'Limpia Piso',
    'Limpia Vidrios':'Limpia Vidrios','Limpia vidrio':'Limpia Vidrios',
    'Lisoform':'Lisoform',
    'Lustra Mueble':'Lustra Muebles','Lustra Muebles':'Lustra Muebles','Lustra muebles':'Lustra Muebles','Lustramueble':'Lustra Muebles',
    'Mopa Avion':'Mopa Avión','Mopa Humeda':'Mopa Húmeda',
    'Mopa Rectangular':'Mopa Rectangular','Mopa Redonda':'Mopa Redonda',
    'Repuesto avion':'Repuesto Mopa Avión',
    'Papel Higienico (UN)':'Papel Higiénico','Papel Interfoliado':'Papel Interfoliado','Papel interfoliado ':'Papel Interfoliado','Papel interfoliado':'Papel Interfoliado',
    'Resma Papel Fotocopia (UN)':'Resma Papel Carta','Resma Papel Fot. Carta (UN)':'Resma Papel Carta',
    'Resma Papel Fot. Oficio (UN)':'Resma Papel Oficio',
    'Poet':'Poet',
    'Removedores':'Revolvedores','Removerdores (UN)':'Revolvedores','Revolvedores':'Revolvedores',
    'Sal':'Sal','Sal Lavavajillas':'Sal Lavavajillas',
    'Servilleta':'Servilletas','Servilletas':'Servilletas',
    'Supremo de Hierbas':'Té de Hierbas','Te (KG)':'Té','Te Lypton (100 unid)':'Té Lipton',
    'Te Natural':'Té de Hierbas','Te Negro (100 unid)':'Té Negro',
    'Te caja amarillo Lipton':'Té Lipton','Te caja negro supremo':'Té Negro',
    'Te caja surtidos':'Té Surtido','Te de Hiervas surtida':'Té de Hierbas','Te de hierbas (100 unid)':'Té de Hierbas',
    'Toalla de Papel Jumbo (UN)':'Toalla de Papel',
    'Vainilla':'Vainilla','Vainilla (KG)':'Vainilla','Vainilla 500gr':'Vainilla',
    'Vasos Polipapel':'Vasos','Vasos de carton':'Vasos',
    'Vinagre':'Vinagre','Vinagre 1 Lt':'Vinagre','Vinagre 1000 ml':'Vinagre','Vinagre 500 ml':'Vinagre',
}
# latin-1 variants
PROD_NORM['Aceto Bals\xe1mico'] = 'Aceto Balsámico'
PROD_NORM['Caf\xe9 instantaneao'] = 'Café Granulado'
PROD_NORM['Caf\xe9 soluble'] = 'Café Granulado'
PROD_NORM['Jab\xf3n de mano'] = 'Jabón de Mano'
PROD_NORM['Leche L\xedquida'] = 'Leche Líquida'
PROD_NORM['Ensencia de Lim\xf3n'] = 'Limón'
PROD_NORM['Mopa Avi\xf3n'] = 'Mopa Avión'
PROD_NORM['Pa\xf1o Esponja'] = 'Paño Esponja'
PROD_NORM['Pa\xf1o Esponja x 6Uni.'] = 'Paño Esponja'
PROD_NORM['Pa\xf1o Verde'] = 'Paño Verde'
PROD_NORM['Pa\xf1os Amarillo'] = 'Paño Amarillo'
PROD_NORM['Pa\xf1os Azul'] = 'Paño Azul'
PROD_NORM['Pa\xf1os Multiuso Amarillo'] = 'Paño Amarillo'
PROD_NORM['Pa\xf1os Multiuso Azul'] = 'Paño Azul'
PROD_NORM['Pa\xf1os Multiuso Verde'] = 'Paño Verde'
PROD_NORM['Trapo Amarillo Uni.'] = 'Paño Amarillo'
PROD_NORM['Trapo Azul Uni.'] = 'Paño Azul'
PROD_NORM['Trapo Verde Uni.'] = 'Paño Verde'
PROD_NORM['T\xe9 Negro'] = 'Té Negro'
PROD_NORM['T\xe9 Premium'] = 'Té de Hierbas'
PROD_NORM['Vasos de cart\xf3n'] = 'Vasos'

PAPELERIA_PRODS = {'Papel Higiénico','Papel Interfoliado','Resma Papel Carta','Resma Papel Oficio','Toalla de Papel'}

def get_categoria(prov, prod):
    if prov == 'Aseo': return 'Aseo'
    if prov in ('Corporate Coffee','Café Caribe'): return 'Cafetería'
    if prod in PAPELERIA_PRODS: return 'Papelería'
    return 'Cafetería'

def get_subcat(prov):
    if prov == 'Corporate Coffee': return 'Corporate Coffee'
    if prov == 'Café Caribe': return 'Café Caribe'
    return None

def fix(s):
    if s is None: return s
    try: return s.encode('latin-1').decode('utf-8')
    except: return s

sede_map = {
    'ABEDULES':'Abedules','APOQUINDO':'Apoquindo','ALTO EL GOLF':'Alto el Golf',
    'LOS MILITARES-BCI':'Los Militares - NACE','MONJITAS':'Monjitas','ISIDORA':'Isidora',
    'VESPUCIO':'Vespucio','NUEVA LAS CONDES':'Nueva Las Condes','KENNEDY':'Kennedy',
    'CERRO EL PLOMO':'Cerro el Plomo','SANTA LUCIA':'Santa Lucia','SANTA ROSA ':'Santa Rosa',
    'PLAZA EGA\xd1A':'Plaza Ega\u00f1a','NEOHAUS':'Neohaus','MANUEL MONTT 2 & 5':'Manuel Montt',
    'SUECIA 7 & 8':'Suecia','FLORIDA CENTER':'Florida Center','TOBALABA 3':'Tobalaba-P3',
}

def parse_qty(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip().lower()
    m = re.search(r'[\d]+(?:[.,]\d+)?', s)
    if m: return float(m.group().replace(',','.'))
    return None

history = defaultdict(list)

for fname in ['docs/Registro Insumos Sedes _ Grupo 1_2025.xlsx','docs/Registro Insumos Sedes _ Grupo 2 _ 2025.xlsx']:
    wb = openpyxl.load_workbook(fname, data_only=True)
    for sheet in wb.sheetnames:
        if 'resumen' in sheet.lower(): continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3: continue
        dates = []
        for cell in rows[1][3:]:
            if hasattr(cell,'strftime'): dates.append(cell.strftime('%Y-%m-%d'))
            elif cell: dates.append(str(cell)[:10])
            else: dates.append(None)
        for row in rows[3:]:
            prov_raw = str(row[1]).strip() if row[1] else None
            prod_raw = str(row[2]).strip() if row[2] else None
            if not prod_raw or prod_raw=='None': continue
            sede_raw = sheet
            sede = sede_map.get(sede_raw, sede_map.get(sede_raw.strip(), fix(sede_raw)))
            prov_fix = fix(prov_raw) if prov_raw else ''
            prod_fix = fix(prod_raw) if prod_raw else ''
            prov = PROV_NORM.get(prov_raw, PROV_NORM.get(prov_fix, prov_fix))
            prod = PROD_NORM.get(prod_raw, PROD_NORM.get(prod_fix, prod_fix))
            key = f"{sede}||{prov}||{prod}"
            for i, cell in enumerate(row[3:]):
                if i >= len(dates) or not dates[i]: continue
                qty = parse_qty(cell)
                if qty is not None:
                    history[key].append((dates[i], qty))

catalog_by_sede = defaultdict(dict)
for key, entries in history.items():
    sede, prov, prod = key.split('||', 2)
    quantities = [q for _,q in entries]
    avg = sum(quantities)/len(quantities)
    min_stock = max(1, round(avg * 0.4))
    cat = get_categoria(prov, prod)
    subcat = get_subcat(prov)
    catalog_by_sede[sede][(prov,prod)] = {
        'proveedor':prov,'producto':prod,'min_stock':min_stock,'categoria':cat,'subcategoria':subcat
    }

latest = {}
trend = {}
for key, entries in history.items():
    s = sorted(entries, key=lambda x: x[0])
    latest[key] = s[-1]
    trend[key] = s[-12:]

print(f"Keys: {len(latest)}, Sedes: {len(catalog_by_sede)}", file=sys.stderr)

# Write catalog
lines = ['// Catalogo normalizado']
lines.append('export const INVENTARIO_CATALOG = {')
for sede in sorted(catalog_by_sede.keys()):
    prods = catalog_by_sede[sede]
    lines.append(f'  "{sede}": [')
    for (prov,prod), p in sorted(prods.items()):
        sub = f', subcategoria: "{p["subcategoria"]}"' if p['subcategoria'] else ''
        lines.append(f'    {{ proveedor: "{prov}", producto: "{prod}", min_stock: {p["min_stock"]}, categoria: "{p["categoria"]}"{sub} }},')
    lines.append('  ],')
lines.append('};')
lines.append(f'export const INVENTARIO_SEDES = {json.dumps(sorted(catalog_by_sede.keys()), ensure_ascii=False)};')

with open('src/inventario_catalog.js','wb') as f:
    f.write('\n'.join(lines).encode('utf-8'))
print("catalog OK", file=sys.stderr)

# Write history
hlines = ['// Historico Excel normalizado']
hlines.append('export const INVENTARIO_EXCEL_LATEST = {')
for key,(fecha,cantidad) in sorted(latest.items()):
    k = key.replace('"','\\"')
    hlines.append(f'  "{k}": {{ fecha: "{fecha}", cantidad: {cantidad} }},')
hlines.append('};')
hlines.append('export const INVENTARIO_EXCEL_TREND = {')
for key,entries in sorted(trend.items()):
    k = key.replace('"','\\"')
    ejs = ', '.join(f'["{f}",{q}]' for f,q in entries)
    hlines.append(f'  "{k}": [{ejs}],')
hlines.append('};')

with open('src/inventario_history.js','wb') as f:
    f.write('\n'.join(hlines).encode('utf-8'))
print("history OK", file=sys.stderr)
